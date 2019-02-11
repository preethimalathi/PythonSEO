from urllib.request import urlopen
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import xlsxwriter
import sqlite3
#READ URL FROM READ.XLSX FILE
g = load_workbook('read.xlsx')
sheet = g.active
url=sheet['A1'].value
#url=input(" Enter a url")
print(url)
u=url
#OPEN THE URL
fo=urlopen(url)
#READ THE URL
a=fo.read()
#print(a)
#ARRANGE IT BY BEAUTIFULSOUP
soup=BeautifulSoup(a,"html.parser")
for script in soup(["script","style"]):
    script.extract()
text=soup.get_text()
lines=(line.strip() for line in text.splitlines())
s=list(lines)
d="".join(s)
c=d.split()
#DISPLAY THE WORD IN LIST FORMAT
#print(c)

#CALCULATE LENGTH OF A WORD
e=len(c)
#CALCULATE THE OCCURENCE
wordfreq = [c.count(w) for w in c]

#READ THE SELECTED WORD FROM READ.XLSX
g = load_workbook('read.xlsx')
sheet = g.active
key=sheet['A2'].value
#key=input(" Enter a keywords")
r=key.split(',')

#CONVERT IT AS DICT FORMAT
dic={}
for s1 in r:
    s1=s1.lower()
    for k,v in zip(c,wordfreq):
        k=k.lower()
        if s1==k:
            dic[s1]=v              
#print(dic)

#CALCULATE DENSITY OF THE WORD            
den=[]
b1=list(dic.values())
c1=list(dic.keys())
length=len(b1)+3

for d1 in b1:
    den.append((d1/e)*100)
    print(den)

#CONNECT TO DATABASE
conn=sqlite3.connect('database7.db')

#conn.execute("create table record1(url varchar(30),keywords varchar(20),density varchar(10))")
#print("Database created successfully");

i=0
try:
    #CREATE THE TABLE
    conn.execute("create table record1(url varchar(30),keywords varchar(20),density varchar(10))")
    while i<=(len(b1)-1):
        #INSERT THE DATA
        conn.execute("insert into record1(url,keywords,density)values(?,?,?)",(u,c1[i],den[i]));
        i=i+1
    #SAVE CHANGES    
    conn.commit()
except BaseException:
    #print("TABLE NOT CREATED")
    #IF THE TABLE IS EXSIST, THEN DROP THE TABLE AND CREATE THE NEW ONE AND INSERT THE DATA
    conn.execute("drop table record1")
    conn.commit()
    #print("Existed Database dropped  successfully");
    conn.execute("create table record1(url varchar(30),keywords varchar(20),density varchar(10))")
    print("Database created successfully");

    i=0
    try:
        while i<=(len(b1)-1):       
            conn.execute("insert into record1(url,keywords,density)values(?,?,?)",(u,c1[i],den[i]));
            i=i+1
        conn.commit()
    except BaseException:
        print("TABLE NOT CREATED")

print("Table inserted successfully");

#FROM THE DATBASE DISPLAY THE CONTENT TO THE FILE
cur=conn.execute("select * from record1")
with open('final_excel_output.csv', 'w') as write_file:
    cursor = conn.cursor()
    for row in cursor.execute('SELECT * FROM record1'):
        write_file.writelines(row)
        print(row)
        for ro in row:
            print(ro)

#CREATE THE WRITE_EXCEL FILE
workbook1=xlsxwriter.Workbook('write_excel.xlsx')
chart = workbook1.add_chart({'type': 'column'})
wsheet = workbook1.add_worksheet()
#ASSIGN THE CELL TO THE VALUE
wsheet.write('A1',u )
wsheet.write_column('A3', c1)
wsheet.write_column('B3', b1)
wsheet.write_column('C3',den)
print("URL,KEYWORDS,DENSITY ARE STORED INTO SPREADSHEET SUCCESSFULLY")
#CREATE THE CHART FOR THAT VALUES
l='=Sheet1!$B$3:$B$'+str(length)
chart.add_series({'values':l})
wsheet.insert_chart('C10', chart)
print("chart drawn into excel successfully")
workbook1.close()


    


